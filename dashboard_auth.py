from __future__ import annotations

import base64
import hashlib
import json
import os
from pathlib import Path

from cryptography.hazmat.primitives.ciphers.aead import AESGCM

PBKDF2_ITERATIONS = 200_000
VALID_ACTIVE = {"active", "فعال"}


def _b64(value: bytes) -> str:
    return base64.b64encode(value).decode("ascii")


def load_users(excel_path: str | Path, dashboard: str = "treasury") -> list[dict[str, str]]:
    from openpyxl import load_workbook

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"فایل کاربران پیدا نشد: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ValueError("فایل کاربران خالی است")

    normalized = {str(value or "").strip().lower(): index for index, value in enumerate(rows[0])}
    aliases = {
        "username": ("username", "user", "نام کاربری", "یوزر"),
        "password": ("password", "pass", "رمز عبور", "پسورد"),
        "treasury": ("treasury", "تریژری", "خزانه"),
    }
    columns: dict[str, int] = {}
    for key, names in aliases.items():
        for name in names:
            if name.lower() in normalized:
                columns[key] = normalized[name.lower()]
                break
        if key not in columns:
            raise ValueError(f"ستون الزامی {key} در فایل کاربران وجود ندارد")

    if dashboard.strip().lower() != "treasury":
        raise ValueError(f"نام داشبورد نامعتبر است: {dashboard}")
    users: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        username = str(row[columns["username"]] or "").strip()
        password = str(row[columns["password"]] or "").strip()
        status = str(row[columns["treasury"]] or "").strip().lower()
        if not username and not password:
            continue
        if not username or not password:
            raise ValueError(f"نام کاربری یا رمز عبور در ردیف {row_number} خالی است")
        folded = username.casefold()
        if folded in seen:
            raise ValueError(f"نام کاربری تکراری در ردیف {row_number}: {username}")
        seen.add(folded)
        if status in VALID_ACTIVE:
            users.append({"username": username, "password": password})
    if not users:
        raise ValueError("هیچ کاربر Active برای داشبورد Treasury تعریف نشده است")
    return users


def protect_html(html: str, excel_path: str | Path, dashboard: str = "treasury", title: str = "داشبورد جریان نقد پوزیترون") -> str:
    users = load_users(excel_path, dashboard)
    data_key = AESGCM.generate_key(bit_length=256)
    content_iv = os.urandom(12)
    content = AESGCM(data_key).encrypt(content_iv, html.encode("utf-8"), None)
    entries = []
    for user in users:
        salt = os.urandom(16)
        iv = os.urandom(12)
        secret = f'{user["username"].casefold()}\0{user["password"]}'.encode("utf-8")
        key = hashlib.pbkdf2_hmac("sha256", secret, salt, PBKDF2_ITERATIONS, 32)
        wrapped_key = AESGCM(key).encrypt(iv, data_key, None)
        entries.append({"u": user["username"].casefold(), "s": _b64(salt), "i": _b64(iv), "k": _b64(wrapped_key)})
    config = json.dumps({"entries": entries, "iv": _b64(content_iv), "ct": _b64(content), "iter": PBKDF2_ITERATIONS}, separators=(",", ":"), ensure_ascii=False)
    safe_title = json.dumps(title, ensure_ascii=False)[1:-1]
    return f'''<!doctype html><html lang="fa" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate"><meta http-equiv="Pragma" content="no-cache"><meta http-equiv="Expires" content="0"><title>{safe_title}</title><style>
*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:grid;place-items:center;background:#07111f;color:#e5edf7;font-family:Vazirmatn,Tahoma,sans-serif}}.box{{width:min(390px,calc(100% - 32px));padding:32px 28px;background:#111d2d;border:1px solid #26364b;border-radius:20px;box-shadow:0 24px 80px #0008}}h1{{font-size:20px;margin:0 0 8px;text-align:center}}p{{color:#9fb0c6;text-align:center;font-size:13px;margin:0 0 24px}}label{{display:block;font-size:12px;color:#b9c7d8;margin:10px 2px 6px}}input{{width:100%;padding:12px 14px;border:1px solid #33465f;border-radius:10px;background:#091524;color:#fff;font-size:14px;outline:none}}input:focus{{border-color:#3b82f6}}button{{width:100%;margin-top:18px;padding:12px;border:0;border-radius:10px;background:#2563eb;color:#fff;font-weight:800;cursor:pointer}}button:disabled{{opacity:.6}}.err{{min-height:20px;margin-top:12px;text-align:center;color:#fb7185;font-size:12px}}
</style></head><body><main class="box"><h1>{safe_title}</h1><p>نام کاربری و رمز عبور خود را وارد کنید</p><form id="login"><label for="username">نام کاربری</label><input id="username" autocomplete="username" autofocus><label for="password">رمز عبور</label><input id="password" type="password" autocomplete="current-password"><button id="submit" type="submit">ورود</button><div id="error" class="err"></div></form></main><script>
const CFG={config};const b64=s=>Uint8Array.from(atob(s),c=>c.charCodeAt(0));const enc=new TextEncoder();login.addEventListener('submit',async e=>{{e.preventDefault();error.textContent='';submit.disabled=true;try{{const u=username.value.trim().toLocaleLowerCase('fa');const item=CFG.entries.find(x=>x.u===u);if(!item)throw 0;const material=await crypto.subtle.importKey('raw',enc.encode(u+'\\0'+password.value),'PBKDF2',false,['deriveKey']);const kek=await crypto.subtle.deriveKey({{name:'PBKDF2',salt:b64(item.s),iterations:CFG.iter,hash:'SHA-256'}},material,{{name:'AES-GCM',length:256}},false,['decrypt']);const dk=await crypto.subtle.decrypt({{name:'AES-GCM',iv:b64(item.i)}},kek,b64(item.k));const contentKey=await crypto.subtle.importKey('raw',dk,{{name:'AES-GCM'}},false,['decrypt']);const plain=await crypto.subtle.decrypt({{name:'AES-GCM',iv:b64(CFG.iv)}},contentKey,b64(CFG.ct));document.open();document.write(new TextDecoder().decode(plain));document.close();}}catch(_){{error.textContent='نام کاربری، رمز عبور یا دسترسی شما صحیح نیست';submit.disabled=false;}}}});
</script></body></html>'''
