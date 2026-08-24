#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_forecast.py — استخراج و پیش‌بینی جریان نقد گروه پوزیترون (Positron TMS)
نسخهٔ v4.2 — بازسازی‌شده از تصمیمات نهایی نشست ۱۴۰۵/۰۵ (مدل تجدید ساده‌شده)

⚠️ این فایل «بازسازی از حافظهٔ گفتگو» است، نه استخراج مستقیم از یک نسخهٔ کارکردهٔ
   قبلی (بر خلاف make_html_v2.py). قبل از تکیه‌کردن کامل روی آن برای گزارش بعدی،
   خروجی‌اش را با گزارش ۱۴۰۵/۰۵/۲۰ (که دستی تأیید شده) مقایسه کن.

=========================== قوانین کلیدی تأییدشده ===========================
۱) تجدید تسهیلات (مدل نهایی، جایگزین همهٔ نسخه‌های قبلی):
   - همهٔ تسهیلات فردای روز کاریِ پرداخت تجدید می‌شوند (اصل)، مگر سررسیدگذشته باشند.
   - هر ردیف/قسط کاملاً مستقل است؛ حتی اگر چند ردیف شماره تسهیلات یکسان داشته باشند،
     هرکدام مدت خودش (افتتاح تا سررسیدِ خودش) را می‌گیرد و زنجیرهٔ تجدید مستقل خودش را
     تا انتهای افق کامل می‌سازد. دوباره‌شماری عمدی است (تصمیم صریح مدیر تأمین مالی).
   - مدت هر چرخه = مدت خودِ همان ردیف: تاریخ افتتاح (دقیقِ همان سررسید در گزارش تسهیلات،
     یا در نبودش نزدیک‌ترین دادهٔ سطح‌کل شماره تسهیلات، یا در نبودِ هردو پیش‌فرض ۳۰ روز)
     تا سررسید. همین مدت تکرار می‌شود تا HORIZON_END، بدون سقف کوتاه‌مدت جداگانه.
۲) Manual Inputs/Policy/Capex Register/Credit Lines از یک فایل *جدا* از SRC خوانده
   می‌شوند (mwb ≠ wb) — چون ذخیرهٔ مجدد SRC با openpyxl مقادیر کش‌شدهٔ فرمول‌های
   شیت‌های اصلی (Daily Cashflow و...) را پاک می‌کند. SRC هرگز نباید دوباره save شود.
۳) ستون‌های مبلغ در Capex Register و Credit Lines به **ریال** هستند (نه میلیون تومان)
   — باید بر MT تقسیم شوند. ستون «باقی‌مانده» ملاک اثر نقدی کپکس است.
۴) دسته‌بندی Manual Inputs: متن دستهٔ فارسی در فایل شامل مقادیری مثل «مالیات» و
   «ارزش افزوده» است که باید به cat داخلی 'tax_ins' نگاشت شوند (نه 'other_out').
   قبلاً این نگاشت ناقص بود و باعث می‌شد مالیات زیر «سایر پرداخت‌ها» برود.
۵) تاریخ‌های مرزی (DAILY_END/WEEKLY_END برای دانه‌بندی نمایش) نسبت به REPORT_DATE
   پویا محاسبه می‌شوند، نه هاردکد — تا با جلوترفتن REPORT_DATE خودکار جلو بروند.
۶) HORIZON_END ثابت است: 1406/03/31.
================================================================================
"""
import os, re, io, json
from collections import defaultdict
import openpyxl

# ---------------- تنظیمات اجرا (برای اپ ویندوز از متغیر محیطی دریافت می‌شود) ----------------
SRC = os.environ.get('TREASURYFLOW_SRC', '/home/claude/decrypted_14050528.xlsx')
REPORT_DATE = os.environ.get('TREASURYFLOW_REPORT_DATE', '1405/05/28')
FACILITIES_LEDGER = os.environ.get('TREASURYFLOW_FACILITIES_LEDGER') or None
MANUAL_INPUTS_FILE = os.environ.get('TREASURYFLOW_MANUAL_INPUTS') or None
SOURCE_PASSWORD = os.environ.get('TREASURYFLOW_PASSWORD', '')
DATA_JSON = os.environ.get('TREASURYFLOW_DATA_JSON', '/home/claude/data.json')
HORIZON_END = os.environ.get('TREASURYFLOW_HORIZON_END', '1406/03/31')
ANNUAL_RATE = float(os.environ.get('TREASURYFLOW_ANNUAL_RATE', '0.23'))
MT = 1e7                     # ۱ میلیون تومان = ۱۰٬۰۰۰٬۰۰۰ ریال
SNAP_DIR = os.environ.get('TREASURYFLOW_SNAPSHOT_DIR', '/home/claude/snapshots')
os.makedirs(SNAP_DIR, exist_ok=True)

# ---------------- تقویم شمسی ----------------
MLEN = {y: [31]*6+[30]*5+[29] for y in range(1400, 1409)}
MLEN[1403] = [31]*6+[30]*5+[30]  # ۱۴۰۳ کبیسه
WDN = ['شنبه','یکشنبه','دوشنبه','سه‌شنبه','چهارشنبه','پنجشنبه','جمعه']
MON = ['','فروردین','اردیبهشت','خرداد','تیر','مرداد','شهریور','مهر','آبان','آذر','دی','بهمن','اسفند']

def parse(ds):
    m = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', str(ds).strip())
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

def daynum(y, m, d):
    n = 0
    for yy in range(1400, y): n += sum(MLEN[yy])
    return n + sum(MLEN[y][:m-1]) + d

def daynum_to_date(nd):
    yy = 1400
    while nd > daynum(yy, 12, MLEN[yy][11]): yy += 1
    base = daynum(yy, 1, 1) - 1; rem = nd - base; mm = 1
    while rem > MLEN[yy][mm-1]: rem -= MLEN[yy][mm-1]; mm += 1
    return yy, mm, rem

def fmt(y, m, d): return f'{y:04d}/{m:02d}/{d:02d}'

def nextday(y, m, d):
    d += 1
    if d > MLEN[y][m-1]: d, m = 1, m+1
    if m > 12: m, y = 1, y+1
    return y, m, d

R0 = daynum(*parse(REPORT_DATE))
ANCHOR = daynum(1405, 4, 17)  # چهارشنبه — لنگر ثابت تقویم کل پروژه
def weekday(y, m, d): return (4 + daynum(y, m, d) - ANCHOR) % 7

HOLIDAYS = {
 '1405/05/22':'اربعین حسینی','1405/05/30':'رحلت پیامبر (ص) و شهادت امام حسن (ع)',
 '1405/06/08':'میلاد رسول اکرم (ص) و امام جعفر صادق (ع)','1405/09/01':'شهادت حضرت فاطمه (س)',
 '1405/09/18':'ولادت حضرت فاطمه (س)','1405/10/10':'ولادت امام علی (ع)',
 '1405/10/24':'مبعث','1405/11/12':'نیمه شعبان','1405/11/22':'پیروزی انقلاب',
 '1405/12/17':'شهادت امام علی (ع)','1405/12/27':'عید فطر','1405/12/28':'تعطیل عید فطر',
 '1405/12/29':'ملی شدن صنعت نفت','1406/01/01':'نوروز','1406/01/02':'نوروز',
 '1406/01/03':'نوروز','1406/01/04':'نوروز','1406/01/12':'روز جمهوری اسلامی',
 '1406/01/13':'روز طبیعت','1406/01/22':'شهادت امام صادق (ع)','1406/03/04':'عید قربان',
 '1406/03/12':'عید غدیر','1406/03/14':'رحلت امام خمینی','1406/03/15':'قیام ۱۵ خرداد',
}

def is_off(y, m, d):
    ds = fmt(y, m, d)
    return weekday(y, m, d) == 6 or ds in HOLIDAYS  # جمعه یا تعطیل رسمی

def next_working(y, m, d):
    while is_off(y, m, d): y, m, d = nextday(y, m, d)
    return y, m, d

# ---------------- بارگذاری فایل خزانه (رمزگشایی‌شده) ----------------
def load_wb(path, password=None):
    if password:
        import msoffcrypto
        with open(path, 'rb') as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            out = io.BytesIO()
            office.decrypt(out)
        return openpyxl.load_workbook(out, data_only=True)
    return openpyxl.load_workbook(path, data_only=True)

wb = load_wb(SRC, SOURCE_PASSWORD or None)

required_source_sheets = {
    'Cash ': 7,
    'Note Receivables': 7,
    'Note Payable': 8,
    'Bounced cheques': 6,
    'Loan ': 15,
}
missing_sheets = [name for name in required_source_sheets if name not in wb.sheetnames]
if missing_sheets:
    raise ValueError('شیت‌های الزامی فایل خزانه پیدا نشد: ' + '، '.join(missing_sheets))
for sheet_name, min_columns in required_source_sheets.items():
    if wb[sheet_name].max_column < min_columns:
        raise ValueError(
            f'ساختار شیت «{sheet_name}» تغییر کرده است؛ حداقل {min_columns} ستون لازم است.'
        )

# فایل Manual Inputs جدا از SRC خوانده می‌شود — SRC هرگز دوباره save نمی‌شود
mwb = wb
if MANUAL_INPUTS_FILE:
    try:
        mwb = openpyxl.load_workbook(MANUAL_INPUTS_FILE, data_only=True)
    except Exception as e:
        print('هشدار: خواندن فایل Manual Inputs ناموفق —', str(e)[:150])

CO_MAP = {  # نگاشت نام فارسی/کلید به کلید داخلی انگلیسی
 'پارس آناهید':'Pars Anahid','گسترش تمدن':'Gostaresh Tamadon','الماس':'Almas',
 'سیمرغ':'Simorgh','رهجو':'Rahjo','کارمان':'karman',
}
CO_FA = {v: k for k, v in CO_MAP.items()}
VALID_CO_NAMES = set(CO_MAP.keys()) | set(CO_MAP.values())  # فایل‌ها گاهی کلید انگلیسی می‌دهند، گاهی فارسی
def co_key(name):
    n = str(name).strip()
    return CO_MAP.get(n, n)

# ---------------- Daily Cashflow: موجودی ابتدای دوره هر شرکت ----------------
# هدر واقعی در ردیف ۴؛ داده از ردیف ۵. ستون ۱۴ = «Available Cash» (مانده قابل برداشت) — همان
# رقمی که با «نقد در دسترس گروه» در داشبورد جمع می‌شود؛ برای هر شرکت روی همه بانک‌هایش جمع می‌زنیم.
# ---------------- Daily Cashflow / Cash: موجودی ابتدای دوره و ریز حساب‌ها ----------------
# منبع درست «نقد در دسترس» ستون Bank Available در شیت Cash است (نه ستون Available Cash
# در Daily Cashflow که شامل خالص‌سازی وام است و رقم متفاوتی می‌دهد). هر دو (moجودی کل و ریز
# حساب‌ها) از همین یک منبع خوانده می‌شوند تا جمع‌شان همیشه هماهنگ بماند.
accounts = []
ws_cash = wb['Cash ']
for r in ws_cash.iter_rows(min_row=6, max_row=ws_cash.max_row, values_only=True):
    if r and r[0] and str(r[0]).strip() in VALID_CO_NAMES and isinstance(r[6], (int, float)):
        accounts.append({'co': co_key(r[0]), 'bank': str(r[1] or '').strip(),
                          'acct': str(r[2] or ''), 'bal': float(r[6] or 0) / MT,
                          'blocked': float(r[5] or 0) / MT if isinstance(r[5], (int, float)) else 0})

opening = defaultdict(float)
for a in accounts:
    opening[a['co']] += a['bal']
opening = dict(opening)

# ---------------- افق نمایش (دانه‌بندی روزانه/هفتگی/ماهانه) — پویا نسبت به REPORT_DATE ----------------
DAILY_END = fmt(*daynum_to_date(R0 + 46))
WEEKLY_END = fmt(*daynum_to_date(R0 + 79))

buckets = []
y, m, d = parse(REPORT_DATE)
while fmt(y, m, d) <= DAILY_END:
    ds = fmt(y, m, d)
    buckets.append({'id': ds, 'type': 'd', 'label': f'{WDN[weekday(y,m,d)]}<br>{m:02d}/{d:02d}',
                     'full': ds, 'off': is_off(y, m, d), 'holiday': HOLIDAYS.get(ds, ''),
                     'start': ds, 'end': ds, 'doff': daynum(y, m, d) - R0})
    y, m, d = nextday(y, m, d)
wn = 1
while fmt(y, m, d) <= WEEKLY_END:
    s = fmt(y, m, d)
    while fmt(y, m, d) < WEEKLY_END and weekday(y, m, d) != 6:
        y, m, d = nextday(y, m, d)
    e = fmt(y, m, d)
    buckets.append({'id': f'w{wn}', 'type': 'w', 'label': f'هفته {wn}<br>{MON[m]}',
                     'full': f'{s} تا {e}', 'off': False, 'holiday': '',
                     'start': s, 'end': e, 'doff': daynum(*parse(s)) - R0})
    wn += 1
    y, m, d = nextday(y, m, d)
    if fmt(y, m, d) >= WEEKLY_END: break
cy, cm = y, m
while fmt(cy, cm, 1) <= HORIZON_END:
    s = fmt(cy, cm, 1); e = fmt(cy, cm, MLEN[cy][cm-1])
    buckets.append({'id': f'm{cy}{cm:02d}', 'type': 'm', 'label': f'{MON[cm]}<br>{cy}',
                     'full': f'{s} تا {e}', 'off': False, 'holiday': '',
                     'start': s, 'end': e, 'doff': daynum(*parse(s)) - R0})
    cm += 1
    if cm > 12: cm = 1; cy += 1

B = buckets
def bucket_of(ds):
    for i, b in enumerate(B):
        if b['start'] <= ds <= b['end']: return i
    return None

# ---------------- ساختار دسته‌ها (دقیقاً منطبق با CATS در make_html_v2 / DATA['cats']) ----------------
CATS = [
 ('nr', 'اسناد دریافتنی', 'in'), ('renew', 'تجدید تسهیلات', 'in'),
 ('dep_release', 'آزادسازی سپرده', 'in'), ('intergroup', 'دریافتی از شرکت‌های گروه', 'in'),
 ('other_in', 'سایر دریافت‌ها', 'in'),
 ('np', 'اسناد پرداختنی', 'out'), ('np_past', 'اسناد پرداختنی سررسیدگذشته', 'out'),
 ('loan', 'تسهیلات پرداختنی', 'out'), ('loan_past', 'تسهیلات پرداختنی سررسیدگذشته', 'out'),
 ('salary', 'حقوق', 'out'), ('tax_ins', 'بیمه و مالیات', 'out'),
 ('loan_fee', 'کارمزد تسهیلات', 'out'), ('ap_petty', 'حساب پرداختنی و تنخواه', 'out'),
 ('other_out', 'سایر پرداخت‌ها', 'out'),
]
STRESS_CATS = {'nr': 'collect', 'renew': 'renew'}

companies = list(opening.keys())
agg = {cid: defaultdict(lambda: [0.0]*len(B)) for cid, _, _ in CATS}
TX = []

def add(cid, co_raw, bucket_idx, amt_rial):
    co = co_key(co_raw)
    if bucket_idx is None: return
    agg[cid][co][bucket_idx] += amt_rial / MT

def tx(ds, co_raw, cid, desc, amt_rial):
    TX.append({'d': ds, 'doff': daynum(*parse(ds)) - R0 if parse(ds) else None,
               'co': co_key(co_raw), 'cat': cid, 'desc': desc, 'amt': amt_rial / MT})

# ---------------- Note Receivables / Note Payable ----------------
# هر دو شیت: هدر در ردیف ۴-۵، داده از ردیف ۶.
# Note Receivables: 0=شرکت 1=بانک 2=شعبه 3=ذینفع 4=شماره چک 5=تاریخ چک 6=مبلغ
# Note Payable:      0=شرکت 1=بانک 2=حساب 3=ذینفع 4=موضوع 5=شماره چک 6=تاریخ چک 7=مبلغ
def read_notes(sheet_name, cid_future, cid_past, is_receivable):
    if sheet_name not in wb.sheetnames: return
    ws_n = wb[sheet_name]
    ben_i, chno_i, ds_i, amt_i = (3, 4, 5, 6) if is_receivable else (3, 5, 6, 7)
    for r in ws_n.iter_rows(min_row=6, max_row=ws_n.max_row, values_only=True):
        co = r[0]
        if not co or str(co).strip() not in VALID_CO_NAMES: continue
        ben, chno, ds, amt = r[ben_i], r[chno_i], r[ds_i], r[amt_i]
        if not ds or not isinstance(amt, (int, float)): continue
        ds_s = str(ds).strip()
        if not parse(ds_s): continue
        sign = 1 if is_receivable else -1
        desc = str(ben or '')
        if ds_s < REPORT_DATE:
            if is_receivable:
                continue  # سررسیدگذشتهٔ دریافتنی -> لیست چک برگشتی/معوق، نه جریان ورودی
            add(cid_past, co, 0, sign*amt); tx(REPORT_DATE, co, cid_past, f'چک معوق {chno} — {desc}', sign*amt)
        else:
            b = bucket_of(ds_s)
            add(cid_future, co, b, sign*amt); tx(ds_s, co, cid_future, f'چک {chno} — {desc}', sign*amt)

read_notes('Note Receivables', 'nr', None, True)
read_notes('Note Payable', 'np', 'np_past', False)

# ---------------- Bounced cheques (چک‌های برگشتی + دریافتنی سررسیدگذشته) ----------------
# هدر در ردیف ۴-۵، داده از ردیف ۶. 0=شرکت 1=بانک 2=ذینفع 3=شماره چک 4=تاریخ چک 5=مبلغ
bounced = []
if 'Bounced cheques' in wb.sheetnames:
    ws_b = wb['Bounced cheques']
    for r in ws_b.iter_rows(min_row=6, max_row=ws_b.max_row, values_only=True):
        if r and r[0] and str(r[0]).strip() in VALID_CO_NAMES and isinstance(r[5], (int, float)):
            bounced.append({'co': co_key(r[0]), 'ben': str(r[2] or ''), 'chno': str(r[3] or ''),
                             'date': str(r[4] or ''), 'amt': float(r[5])/MT, 'src': 'چک برگشتی'})
if 'Note Receivables' in wb.sheetnames:
    ws_nr = wb['Note Receivables']
    for r in ws_nr.iter_rows(min_row=6, max_row=ws_nr.max_row, values_only=True):
        co = r[0]
        if not co or str(co).strip() not in VALID_CO_NAMES: continue
        ben, chno, ds, amt = r[3], r[4], r[5], r[6]
        if ds and isinstance(amt, (int, float)) and str(ds).strip() < REPORT_DATE:
            bounced.append({'co': co_key(co), 'ben': str(ben or ''), 'chno': str(chno or ''),
                             'date': str(ds), 'amt': float(amt)/MT, 'src': 'معوق اسناد دریافتنی'})

# ---------------- Facilities ledger — تاریخ افتتاح هر ردیف/شماره ----------------
orig_date_by_no = {}
orig_date_by_no_due = {}
if FACILITIES_LEDGER:
    try:
        lwb = openpyxl.load_workbook(FACILITIES_LEDGER, data_only=True)
        ledger_sheet_by_trimmed_name = {name.strip(): name for name in lwb.sheetnames}
        for sheet_name, no_col, orig_col, due_col in [
            (' تسهیلات فعال ', 3, 7, 8), (' تسهیلات تسویه شده   ', 3, 6, 7)]:
            actual_sheet_name = sheet_name if sheet_name in lwb.sheetnames else ledger_sheet_by_trimmed_name.get(sheet_name.strip())
            if not actual_sheet_name: continue
            ws_l = lwb[actual_sheet_name]
            for r in ws_l.iter_rows(min_row=3, max_row=ws_l.max_row, values_only=True):
                no, orig, due = r[no_col], r[orig_col], r[due_col]
                if no and due and parse(due) and orig and parse(orig):
                    no_s = str(no).strip(); due_s = str(due).strip(); od = str(orig).strip()
                    if no_s not in orig_date_by_no or od < orig_date_by_no[no_s]:
                        orig_date_by_no[no_s] = od
                    orig_date_by_no_due[(no_s, due_s)] = od
    except Exception as e:
        print('هشدار: خواندن گزارش تسهیلات ناموفق —', str(e)[:150])
print(f'لجر تسهیلات: {len(orig_date_by_no_due)} ردیف با تاریخ افتتاح دقیق، '
      f'{len(orig_date_by_no)} شماره با حداقل یک تاریخ افتتاح شناخته‌شده.')

# ---------------- Loans: منطق تجدید نهایی (هر ردیف مستقل، مدت=مدت خودش) ----------------
loans, raw_rows, projected = [], [], []
ws = wb['Loan ']
for r in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
    co, bank, acct, ltype = r[0], r[1], r[2], r[3]
    loanno, dpay = r[5], r[8]
    prin, intr, total = r[11], r[12], r[14]
    if not co or not isinstance(total, (int, float)) or not dpay: continue
    ds = str(dpay).strip(); p = parse(ds)
    if not p: continue
    raw_rows.append({'co': co, 'bank': str(bank or '').strip(), 'ltype': str(ltype or '').strip(),
                      'no': str(loanno or ''), 'due': ds, 'prin': prin or 0, 'intr': intr or 0, 'total': total})

MAX_CYCLES = 400
for rr in raw_rows:
    co, bank, ltype, loanno, ds = rr['co'], rr['bank'], rr['ltype'], rr['no'], rr['due']
    prin, intr, total = rr['prin'], rr['intr'], rr['total']
    rec = {'co': co_key(co), 'bank': bank, 'type': ltype, 'no': loanno,
           'due': ds, 'total': total/MT, 'prin': (prin or 0)/MT, 'intr': (intr or 0)/MT}
    if ds < REPORT_DATE:
        add('loan_past', co, 0, -total)
        tx(REPORT_DATE, co, 'loan_past', f'{ltype} {bank} معوق (سررسید {ds})', -total)
        rec.update({'status': 'past', 'pay': REPORT_DATE, 'renew': '', 'renew_amt': 0})
        loans.append(rec); continue

    p = parse(ds)
    py, pm, pd_ = next_working(*p); pay_ds = fmt(py, pm, pd_)
    add('loan', co, bucket_of(pay_ds), -total)
    ry, rm, rd = next_working(*nextday(py, pm, pd_)); renew_ds = fmt(ry, rm, rd)
    add('renew', co, bucket_of(renew_ds), (prin or 0))
    tx(pay_ds, co, 'loan', f'{ltype} {bank} قسط', -total)
    tx(renew_ds, co, 'renew', f'تجدید {ltype} {bank}', (prin or 0))
    rec.update({'status': 'due', 'pay': pay_ds, 'renew': renew_ds, 'renew_amt': (prin or 0)/MT})
    loans.append(rec)

    if not prin: continue
    own_orig = orig_date_by_no_due.get((loanno, ds)) or orig_date_by_no.get(loanno)
    if own_orig and parse(own_orig):
        duration = daynum(*parse(ds)) - daynum(*parse(own_orig))
        dur_src = ('دهانهٔ خودِ همین ردیف (افتتاح تا سررسید، گزارش تسهیلات)'
                   if orig_date_by_no_due.get((loanno, ds))
                   else 'افتتاح سطح کل شماره تسهیلات (نزدیک‌ترین دادهٔ موجود)')
        dur_conf = 90 if orig_date_by_no_due.get((loanno, ds)) else 70
    else:
        duration, dur_src, dur_conf = 30, 'بدون دادهٔ افتتاح در گزارش تسهیلات — پیش‌فرض ۳۰ روز', 30
    if duration < 1:
        duration, dur_src, dur_conf = 30, 'دهانهٔ نامعتبر (≤۰ روز) — پیش‌فرض ۳۰ روز', 30

    cyc_prin = prin; cur_y, cur_m, cur_d = ry, rm, rd
    for cyc in range(1, MAX_CYCLES+1):
        nd = daynum(cur_y, cur_m, cur_d) + duration
        yy, mm, dd = daynum_to_date(nd)
        cy, cm, cdd = next_working(yy, mm, dd)
        cyc_due = fmt(cy, cm, cdd)
        if cyc_due > HORIZON_END: break
        b = bucket_of(cyc_due)
        cyc_interest = cyc_prin * ANNUAL_RATE * (duration / 365)
        cyc_total = cyc_prin + cyc_interest
        add('loan', co, b, -cyc_total)
        tx(cyc_due, co, 'loan', f'{ltype} {bank} چرخه بعدی #{cyc} (برآوردی)', -cyc_total)
        ry2, rm2, rd2 = next_working(*nextday(cy, cm, cdd)); renew2_ds = fmt(ry2, rm2, rd2)
        add('renew', co, bucket_of(renew2_ds), cyc_prin)
        tx(renew2_ds, co, 'renew', f'تجدید {ltype} {bank} چرخه #{cyc} (برآوردی)', cyc_prin)
        projected.append({'co': co_key(co), 'bank': bank, 'type': ltype, 'no': loanno, 'cycle': cyc,
                           'due': cyc_due, 'renew': renew2_ds, 'amt': round(cyc_prin/MT, 1),
                           'interest': round(cyc_interest/MT, 1), 'total': round(cyc_total/MT, 1),
                           'tenor': duration, 'tenor_src': dur_src, 'tenor_conf': dur_conf, 'from_due': ds})
        cur_y, cur_m, cur_d = ry2, rm2, rd2

# ---------------- Timeline (برای تب خط زمانی تسهیلات) ----------------
timeline = []
loans_by_no = defaultdict(list)
for l in loans:
    if l['status'] == 'due': loans_by_no[l['no']].append(l)
for no, rows in loans_by_no.items():
    rows.sort(key=lambda x: x['due'])
    fallback_start = orig_date_by_no.get(no)
    for idx, l in enumerate(rows):
        own_start = orig_date_by_no_due.get((no, l['due'])) or fallback_start
        renew_end = rows[idx+1]['due'] if idx+1 < len(rows) else l['renew']
        timeline.append({'no': no, 'co': l['co'], 'bank': l['bank'], 'type': l['type'],
                          'start': own_start, 'end': l['due'], 'amt': l['prin'], 'kind': 'real',
                          'origin_due': l['due'], 'renew_date': l['renew'], 'renew_amt': l['renew_amt'],
                          'renew_end': renew_end})
        fallback_start = l['renew']
        prev_renew = l['renew']
        cyc_list = sorted([x for x in projected if x['no'] == no and x['from_due'] == l['due']],
                           key=lambda x: x['due'])
        for cidx, p in enumerate(cyc_list):
            cyc_renew_end = cyc_list[cidx+1]['due'] if cidx+1 < len(cyc_list) else p['renew']
            timeline.append({'no': no, 'co': p['co'], 'bank': p['bank'], 'type': p['type'],
                              'start': prev_renew, 'end': p['due'], 'amt': p['amt'], 'kind': 'projected',
                              'origin_due': l['due'], 'renew_date': p['renew'], 'renew_amt': p['amt'],
                              'renew_end': cyc_renew_end})
            prev_renew = p['renew']
for l in loans:
    if l['status'] == 'past':
        timeline.append({'no': l['no'], 'co': l['co'], 'bank': l['bank'], 'type': l['type'],
                          'start': orig_date_by_no_due.get((l['no'], l['due'])) or orig_date_by_no.get(l['no']),
                          'end': REPORT_DATE, 'origin_due': l['due'], 'amt': l['total'], 'kind': 'past_due'})

# ---------------- Manual Inputs (شیت جدا، فایل جدا) ----------------
# نگاشت دستهٔ فارسیِ ستون «cat» فایل Manual Inputs به cid داخلی — شامل واریانت‌های واقعی متن
MANUAL_CAT = {
 'حقوق': 'salary', 'بیمه و مالیات': 'tax_ins', 'مالیات': 'tax_ins', 'ارزش افزوده': 'tax_ins',
 'کارمزد تسهیلات': 'loan_fee', 'حساب پرداختنی و تنخواه': 'ap_petty', 'تنخواه': 'ap_petty',
 'آزادسازی سپرده': 'dep_release', 'دریافتی از شرکت‌های گروه': 'intergroup',
}
def manual_cat_of(raw_cat, amt):
    raw = str(raw_cat or '').strip()
    for key, cid in MANUAL_CAT.items():
        if key in raw:
            return cid
    return 'other_in' if amt >= 0 else 'other_out'

manual_display = []
if 'Manual Inputs' in mwb.sheetnames:
    for r in mwb['Manual Inputs'].iter_rows(min_row=2, values_only=True):
        ds, co, direction, cat_raw, desc, amt = r[0], r[1], r[2], r[3], r[4], r[5]
        if not ds or not co or not amt: continue
        ds_s = str(ds).strip()
        if not parse(ds_s): continue
        dir_s = str(direction or '').strip()
        # فایل Manual Inputs مبلغ را همیشه به‌صورت قدرمطلق ثبت می‌کند؛ علامت باید از ستون «جهت» اعمال شود
        if 'پرداخت' in dir_s: amt = -abs(amt)
        elif 'دریافت' in dir_s: amt = abs(amt)
        cid = manual_cat_of(cat_raw, amt)
        manual_display.append({'date': ds_s, 'co': co_key(co), 'dir': str(direction or ''),
                                'cat': str(cat_raw or ''), 'desc': str(desc or ''), 'amt': amt/MT})
        b = bucket_of(ds_s)
        add(cid, co, b, amt)
        tx(ds_s, co, cid, str(desc or cat_raw or ''), amt)

# ---------------- Policy (حداقل نقد مصوب) ----------------
policy = {c: {'min': 0.0, 'line': 0.0} for c in companies}
if 'Policy' in mwb.sheetnames:
    for r in mwb['Policy'].iter_rows(min_row=2, values_only=True):
        if r[0] and co_key(r[0]) in policy:
            policy[co_key(r[0])]['min'] = float(r[1] or 0)
            policy[co_key(r[0])]['line'] = float(r[2] or 0)

# ---------------- Capex Register (ستون‌های مبلغ به ریال) ----------------
capex = []
if 'Capex Register' in mwb.sheetnames:
    for r in mwb['Capex Register'].iter_rows(min_row=2, values_only=True):
        if r[1] and r[2]:
            need13w = float(r[7])/MT if r[7] not in (None, '') else None
            remaining = float(r[6])/MT if r[6] not in (None, '') else 0
            cash_effect = need13w if need13w is not None else remaining
            capex.append({'start': str(r[0] or ''), 'co': co_key(r[1]), 'project': str(r[2] or ''),
                           'cat': str(r[3] or ''), 'budget': float(r[4])/MT if r[4] not in (None,'') else 0,
                           'spent': float(r[5])/MT if r[5] not in (None,'') else 0, 'remaining': remaining,
                           'need13w': cash_effect, 'funding': str(r[8] or ''), 'payback': r[9], 'irr': r[10],
                           'stage': str(r[11] or ''), 'status': str(r[12] or ''), 'owner': str(r[13] or ''),
                           'notes': str(r[14] or '')})
            if r[0] and cash_effect:
                ds = str(r[0]).strip()
                if parse(ds) and ds >= REPORT_DATE:
                    b = bucket_of(ds)
                    if b is not None:
                        add('other_out', r[1], b, -cash_effect*MT)
                        tx(ds, r[1], 'other_out', f'کپکس: {r[2]}', -cash_effect*MT)

# ---------------- Sales Budget (فقط برای پیش‌فرض ردیف «فروش (دستی)» در تب سر به سر — در هیچ محاسبهٔ نقدی دیگری استفاده نمی‌شود) ----------------
SALES_CO_MAP = {'سیمرغ': 'Simorgh', 'الماس': 'Almas', 'گسترش': 'Gostaresh Tamadon',
                'پارس آناهید': 'Pars Anahid', 'رهجو': 'Rahjo'}
sales_budget = {}
rep_year = int(REPORT_DATE[:4])
if 'Sales Budget' in mwb.sheetnames:
    sws = mwb['Sales Budget']
    header = [str(c or '').strip() for c in next(sws.iter_rows(min_row=1, max_row=1, values_only=True))]
    co_cols = {ci: SALES_CO_MAP[h] for ci, h in enumerate(header) if h in SALES_CO_MAP}
    for r in sws.iter_rows(min_row=2, values_only=True):
        month_no = r[0]
        if not isinstance(month_no, (int, float)) or not (1 <= int(month_no) <= 12):
            continue
        rep_month = int(REPORT_DATE[5:7])
        yy = rep_year + 1 if int(month_no) < rep_month else rep_year
        ym = f'{yy}/{int(month_no):02d}'
        sales_budget[ym] = {}
        for ci, cok in co_cols.items():
            v = r[ci]
            if v not in (None, ''):
                sales_budget[ym][cok] = round(float(v) * 1000, 1)  # میلیارد تومان → میلیون تومان

# ---------------- Credit Lines (ستون‌های مبلغ به ریال) ----------------
credit_lines = []
if 'Credit Lines' in mwb.sheetnames:
    for r in mwb['Credit Lines'].iter_rows(min_row=2, values_only=True):
        if r[0] and r[3]:
            limit_ = float(r[3])/MT; drawn_ = float(r[4])/MT if r[4] not in (None,'') else 0
            credit_lines.append({'co': co_key(r[0]), 'bank': str(r[1] or ''), 'type': str(r[2] or ''),
                                  'limit': limit_, 'drawn': drawn_, 'available': limit_-drawn_,
                                  'maturity': str(r[6] or ''), 'rate': r[7], 'covenant': str(r[8] or ''),
                                  'collateral': str(r[9] or ''), 'owner': str(r[10] or ''), 'notes': str(r[11] or '')})

# ---------------- Assumptions ledger ----------------
ASSUMPTIONS = [
 {'id':'A1','area':'اسناد پرداختنی/دریافتنی','rule':'سررسید مصادف با جمعه/تعطیل رسمی به نخستین روز کاری بعد منتقل می‌شود.','param':'—','conf':100},
 {'id':'A2','area':'اسناد پرداختنی سررسیدگذشته','rule':'چک پرداختنی با تاریخ قبل از روز گزارش، به‌صورت کامل در روز جاری تعهد فرض می‌شود.','param':'۱۰۰٪ در روز جاری','conf':100},
 {'id':'A3','area':'اسناد دریافتنی سررسیدگذشته','rule':'به لیست چک‌های برگشتی/معوق منتقل و از پیش‌بینی ورودی حذف می‌شود.','param':'۰٪ در پیش‌بینی پایه','conf':100},
 {'id':'A4','area':'اسناد دریافتنی آتی','rule':'در تاریخ سررسید (روز کاری) به‌طور کامل وصول فرض می‌شود.','param':'۱۰۰٪ وصول — قابل استرس‌تست','conf':85},
 {'id':'A5','area':'تسهیلات — پرداخت','rule':'هر قسط به میزان مبلغ کل (Total amount) در سررسید پرداخت می‌شود.','param':'۱۰۰٪ از Total amount','conf':100},
 {'id':'A6','area':'تسهیلات — تجدید','rule':'همهٔ تسهیلات فردای روز کاریِ پرداخت تجدید می‌شوند (اصل)، مگر سررسیدگذشته باشند.','param':'۱۰۰٪ اصل، T+۱ روز کاری','conf':100},
 {'id':'A7','area':'تسهیلات — چرخه بعدی','rule':'هر ردیف/قسط مستقل زنجیرهٔ تجدید خودش را می‌گیرد؛ مدت هر چرخه = مدت خودِ همان ردیف، تا انتهای افق کامل تکرار می‌شود.','param':'مدت به تفکیک هر ردیف','conf':80},
 {'id':'A8','area':'کپکس و سرمایه‌گذاری','rule':'ستون «باقی‌مانده» شیت Capex Register در تاریخ شروع پروژه به‌عنوان خروجی نقد لحاظ می‌شود.','param':'از Manual Inputs','conf':70},
 {'id':'A9','area':'خط اعتباری در برابر مانده','rule':'صرفاً نمایشی است؛ روی مدل نقد اثر مستقیم ندارد.','param':'—','conf':None},
 {'id':'A10','area':'بهرهٔ چرخه‌های برآوردی','rule':'برای چرخه‌های برآوردی (که در فایل روزانه هنوز قسط واقعی ندارند)، بهره با نرخ سالانهٔ تخمینی ۲۳٪ محاسبه می‌شود: اصل × ۲۳٪ × (مدت چرخه/۳۶۵). این نرخ واقعی هر تسهیلات نیست، فقط میانگین بازار برای برآورد نظم‌بخشیدن به هزینهٔ آیندهٔ تجدید تسهیلات.','param':'۲۳٪ سالانه — قابل تغییر','conf':50},
]

# ---------------- Diff نسبت به آخرین اسنپ‌شات ----------------
# نکته مهم: main.js چک می‌کند `if(!DATA.diff)` برای پیام «اولین اسنپ‌شات» — پس وقتی اسنپ‌شات
# قبلی نداریم باید diff=None (نه دیکشنری با available:False) باشد. وقتی موجود است، باید شامل
# رکوردهای *کامل* وام/چک (نه فقط شناسه) باشد چون جدول‌های نمایش مستقیم از این فیلدها می‌خوانند:
# co, bank, type, due, total (برای وام) و co, ben, chno, amt (برای چک).
def snapshot_key(ds): return ds.replace('/', '-')
snap_files = sorted(f for f in os.listdir(SNAP_DIR) if f.endswith('.json') and f < snapshot_key(REPORT_DATE)+'.json')
diff = None
if snap_files:
    prev = json.load(open(os.path.join(SNAP_DIR, snap_files[-1]), encoding='utf-8'))
    prev_date = snap_files[-1].replace('.json','').replace('-','/')
    open_now = sum(opening.values()); open_prev = sum(prev.get('opening', {}).values())
    open_co_delta = {c: round(opening.get(c,0.0) - prev.get('opening',{}).get(c,0.0), 1) for c in companies}
    np_past_now = -sum(a[0] for a in agg['np_past'].values()) if agg['np_past'] else 0
    np_past_prev = -sum(a[0] for a in prev.get('agg',{}).get('np_past',{}).values()) if prev.get('agg',{}).get('np_past') else 0
    loan_past_now = -sum(a[0] for a in agg['loan_past'].values()) if agg['loan_past'] else 0
    loan_past_prev = -sum(a[0] for a in prev.get('agg',{}).get('loan_past',{}).values()) if prev.get('agg',{}).get('loan_past') else 0
    bounced_now = sum(x['amt'] for x in bounced); bounced_prev = sum(x['amt'] for x in prev.get('bounced',[]))
    nr_now = sum(sum(a) for a in agg['nr'].values()) if agg['nr'] else 0
    nr_prev = sum(sum(a) for a in prev.get('agg',{}).get('nr',{}).values()) if prev.get('agg',{}).get('nr') else 0

    loans_by_key_now = {(l['no'], l['due']): l for l in loans if l['status'] == 'past'}
    loans_by_key_prev = {(l['no'], l['due']): l for l in prev.get('loans', []) if l['status'] == 'past'}
    newly_past_loans = [loans_by_key_now[k] for k in loans_by_key_now.keys() - loans_by_key_prev.keys()]
    cleared_past_loans = [loans_by_key_prev[k] for k in loans_by_key_prev.keys() - loans_by_key_now.keys()]

    bnc_by_key_now = {(b['co'], b['chno']): b for b in bounced}
    bnc_by_key_prev = {(b['co'], b['chno']): b for b in prev.get('bounced', [])}
    new_bounced = [bnc_by_key_now[k] for k in bnc_by_key_now.keys() - bnc_by_key_prev.keys()]
    cleared_bounced = [bnc_by_key_prev[k] for k in bnc_by_key_prev.keys() - bnc_by_key_now.keys()]

    diff = {'available': True, 'prev_date': prev_date,
            'open_delta': round(open_now-open_prev,1), 'open_co_delta': open_co_delta,
            'np_past_delta': round(np_past_now-np_past_prev,1),
            'loan_past_delta': round(loan_past_now-loan_past_prev,1), 'bounced_delta': round(bounced_now-bounced_prev,1),
            'nr_delta': round(nr_now-nr_prev,1),
            'newly_past_loans': newly_past_loans, 'cleared_past_loans': cleared_past_loans,
            'new_bounced': new_bounced, 'cleared_bounced': cleared_bounced}

# ---------------- سریالایز خروجی ----------------
DATA = {
 'report_date': REPORT_DATE, 'version': f'v4.2 | داده تا {REPORT_DATE}',
 'buckets': B, 'companies': [{'key': c, 'fa': CO_FA.get(c, c)} for c in companies],
 'cats': [{'id': c[0], 'label': c[1], 'dir': c[2]} for c in CATS],
 'stress_cats': STRESS_CATS,
 'opening': {c: round(opening.get(c,0.0),1) for c in companies},
 'agg': {cid: {co: [round(v,1) for v in arr] for co, arr in m.items()} for cid, m in agg.items()},
 'loans': loans, 'bounced': bounced, 'vaset': [], 'manual': manual_display,
 'holidays': HOLIDAYS, 'accounts': accounts, 'tx': TX, 'projected': projected, 'timeline': timeline,
 'policy': policy, 'assumptions': ASSUMPTIONS, 'diff': diff,
 'capex': capex, 'credit_lines': credit_lines, 'sales_budget': sales_budget,
}
data_parent = os.path.dirname(os.path.abspath(DATA_JSON))
os.makedirs(data_parent, exist_ok=True)
with open(DATA_JSON, 'w', encoding='utf-8') as f:
    json.dump(DATA, f, ensure_ascii=False)
with open(os.path.join(SNAP_DIR, snapshot_key(REPORT_DATE)+'.json'), 'w', encoding='utf-8') as f:
    json.dump(DATA, f, ensure_ascii=False)

print(f'buckets: {len(B)} | loans: {len(loans)} | bounced: {len(bounced)} | manual: {len(manual_display)} | '
      f'tx: {len(TX)} | projected: {len(projected)}')
print('opening (m.toman):', {c: round(v) for c,v in opening.items()}, '| sum:', round(sum(opening.values())))
print('diff available:', diff is not None, ('vs '+diff.get('prev_date','')) if diff else '')
