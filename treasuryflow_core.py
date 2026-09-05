from __future__ import annotations

from contextlib import redirect_stderr, redirect_stdout
from dataclasses import asdict, dataclass
from datetime import datetime
import hashlib
import importlib.util
import io
import json
import os
from pathlib import Path
import re
import runpy
import shutil
import subprocess
import sys
import time
from typing import Callable, Iterable
import uuid


APP_VERSION = "1.0.0"
DEFAULT_PASSWORD = "1123"
DEFAULT_HORIZON = "1406/03/31"
STATE_FILE = ".treasuryflow_state.json"
SIGNATURE_FILE = "_last_processed_signature.txt"
LOG_FILE = "treasuryflow.log"
UPLOAD_REQUEST_FILE = "upload_request.json"
ACCESS_FILE_ENV = "TREASURYFLOW_ACCESS_FILE"
ACCESS_DASHBOARD = "treasury"
ACCESS_TITLE = "داشبورد جریان نقد پوزیترون"
PUBLISH_REPO_ENV = "TREASURYFLOW_PUBLISH_REPO"

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
DATE_RE = re.compile(r"(?<!\d)(14\d{2})[-_/](\d{1,2})[-_/](\d{1,2})(?!\d)")


class TreasuryFlowError(RuntimeError):
    """A user-facing processing error."""


@dataclass(frozen=True)
class SourceSet:
    daily: Path | None
    manual: Path | None
    facilities: Path | None

    def existing(self) -> list[Path]:
        return [path for path in (self.daily, self.manual, self.facilities) if path]


@dataclass(frozen=True)
class ProcessResult:
    report_date: str
    report_path: Path
    index_path: Path
    data_path: Path
    signature: str
    sources: SourceSet


def resource_root() -> Path:
    frozen_root = getattr(sys, "_MEIPASS", None)
    return Path(frozen_root) if frozen_root else Path(__file__).resolve().parent


def application_home() -> Path:
    configured = os.environ.get("TREASURYFLOW_HOME")
    if configured:
        return Path(configured).expanduser().resolve()
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def access_file() -> Path:
    configured = os.environ.get(ACCESS_FILE_ENV)
    if configured:
        return Path(configured).expanduser().resolve()
    return Path.home() / "Desktop" / "DashboardAccess" / "dashboard_users.xlsx"


def normalize_text(value: object) -> str:
    return str(value or "").translate(PERSIAN_DIGITS).translate(ARABIC_DIGITS).strip()


def normalize_date(value: object) -> str | None:
    match = DATE_RE.search(normalize_text(value))
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    return f"{year:04d}/{month:02d}/{day:02d}"


def _filename_kind(path: Path) -> str | None:
    name = normalize_text(path.stem).lower().replace("_", " ").replace("-", " ")
    if "template" in name or "نمونه" in name:
        return None
    if "manual input" in name or "ورودی دستی" in name:
        return "manual"
    if "گزارش تسهیلات" in name or "facilities" in name or "loan ledger" in name:
        return "facilities"
    if "treasury daily" in name or name == "daily" or "گزارش روزانه خزانه" in name:
        return "daily"
    return None


def _load_workbook(path: Path, password: str = DEFAULT_PASSWORD):
    import openpyxl

    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as direct_error:
        try:
            import msoffcrypto

            decrypted = io.BytesIO()
            with path.open("rb") as handle:
                office = msoffcrypto.OfficeFile(handle)
                office.load_key(password=password)
                office.decrypt(decrypted)
            decrypted.seek(0)
            return openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
        except Exception as encrypted_error:
            raise TreasuryFlowError(
                f"فایل «{path.name}» قابل خواندن نیست. رمز یا ساختار فایل را بررسی کنید."
            ) from encrypted_error


def classify_path(path: Path, password: str = DEFAULT_PASSWORD) -> str:
    by_name = _filename_kind(path)
    if by_name:
        return by_name
    workbook = _load_workbook(path, password)
    try:
        names = set(workbook.sheetnames)
        trimmed = {name.strip() for name in names}
        if {"Cash ", "Note Receivables", "Note Payable", "Loan "}.issubset(names):
            return "daily"
        if "Manual Inputs" in names and ("Policy" in names or "Sales Budget" in names):
            return "manual"
        if {"تسهیلات فعال", "تسهیلات تسویه شده"} & trimmed:
            return "facilities"
    finally:
        workbook.close()
    raise TreasuryFlowError(f"نوع فایل «{path.name}» شناخته نشد.")


def detect_report_date(path: Path, password: str = DEFAULT_PASSWORD) -> str:
    from_name = normalize_date(path.name)
    if from_name:
        return from_name
    workbook = _load_workbook(path, password)
    try:
        if "Daily Cashflow" in workbook.sheetnames:
            for cell in ("B2", "A2", "B1", "A1"):
                detected = normalize_date(workbook["Daily Cashflow"][cell].value)
                if detected:
                    return detected
    finally:
        workbook.close()
    raise TreasuryFlowError(
        "تاریخ گزارش از نام فایل یا شیت Daily Cashflow پیدا نشد. "
        "نام فایل را مانند Treasury Daily Report - 1405-05-28.xlsx بگذارید."
    )


def _latest(paths: Iterable[Path], *, daily: bool = False) -> Path | None:
    candidates = list(paths)
    if not candidates:
        return None
    if daily:
        def daily_key(path: Path) -> tuple[str, int]:
            return (normalize_date(path.name) or "0000/00/00", path.stat().st_mtime_ns)
        return max(candidates, key=daily_key)
    return max(candidates, key=lambda path: path.stat().st_mtime_ns)


def find_sources(home: Path) -> SourceSet:
    home = Path(home)
    buckets: dict[str, list[Path]] = {"daily": [], "manual": [], "facilities": []}
    for path in home.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        kind = _filename_kind(path)
        if kind:
            buckets[kind].append(path)
    return SourceSet(
        daily=_latest(buckets["daily"], daily=True),
        manual=_latest(buckets["manual"]),
        facilities=_latest(buckets["facilities"]),
    )


def file_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()[:20]


def source_signature(sources: SourceSet, users_path: Path | None = None) -> str:
    parts: list[str] = []
    for key, path in (("D", sources.daily), ("M", sources.manual), ("F", sources.facilities)):
        if path:
            stat = path.stat()
            parts.append(f"{key}:{path.name}:{stat.st_size}:{stat.st_mtime_ns}:{file_digest(path)}")
        else:
            parts.append(f"{key}:-")
    if users_path:
        users_path = Path(users_path)
        if users_path.exists():
            stat = users_path.stat()
            parts.append(f"A:{users_path.name}:{stat.st_size}:{stat.st_mtime_ns}:{file_digest(users_path)}")
        else:
            parts.append(f"A:{users_path}:-")
    return "|".join(parts)


def wait_until_stable(paths: Iterable[Path], delay: float = 1.0) -> None:
    paths = list(paths)
    before = [(path.stat().st_size, path.stat().st_mtime_ns) for path in paths]
    time.sleep(delay)
    after = [(path.stat().st_size, path.stat().st_mtime_ns) for path in paths]
    if before != after:
        raise TreasuryFlowError("کپی فایل هنوز کامل نشده است؛ چند ثانیه بعد دوباره تلاش می‌شود.")


def load_state(home: Path) -> dict:
    path = Path(home) / STATE_FILE
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _atomic_text(path: Path, text: str) -> None:
    pending = path.with_name(path.name + ".pending")
    pending.write_text(text, encoding="utf-8")
    os.replace(pending, path)


def save_state(home: Path, result: ProcessResult) -> None:
    payload = {
        "version": APP_VERSION,
        "last_success": datetime.now().astimezone().isoformat(timespec="seconds"),
        "report_date": result.report_date,
        "signature": result.signature,
        "report": result.report_path.name,
        "sources": {
            key: str(value) if value else None
            for key, value in asdict(result.sources).items()
        },
    }
    _atomic_text(Path(home) / STATE_FILE, json.dumps(payload, ensure_ascii=False, indent=2))
    _atomic_text(Path(home) / SIGNATURE_FILE, result.signature + "\n")


def log_message(home: Path, message: str) -> None:
    timestamp = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")
    with (Path(home) / LOG_FILE).open("a", encoding="utf-8") as handle:
        handle.write(f"[{timestamp}] {message.rstrip()}\n")


def upload_request_path(home: Path) -> Path:
    return Path(home) / ".treasuryflow" / UPLOAD_REQUEST_FILE


def mark_upload_request(home: Path, files: Iterable[Path]) -> Path:
    path = upload_request_path(home)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "id": uuid.uuid4().hex,
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "files": [Path(item).name for item in files],
    }
    _atomic_text(path, json.dumps(payload, ensure_ascii=False, indent=2))
    return path


def has_upload_request(home: Path) -> bool:
    return upload_request_path(home).exists()


def clear_upload_request(home: Path) -> None:
    upload_request_path(home).unlink(missing_ok=True)


def has_changed(home: Path, sources: SourceSet | None = None) -> bool:
    sources = sources or find_sources(home)
    if not sources.daily:
        return False
    return load_state(home).get("signature") != source_signature(sources, access_file())


def _load_html_builder():
    module_path = resource_root() / "pipeline" / "make_html_v2.py"
    spec = importlib.util.spec_from_file_location("treasuryflow_html_builder", module_path)
    if spec is None or spec.loader is None:
        raise TreasuryFlowError("سازنده داشبورد در بسته برنامه پیدا نشد.")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.build_html


def _protect_html(html: str, users_path: Path) -> str:
    try:
        from dashboard_auth import protect_html
        # Shared trial key is added only in memory immediately before encryption.
        # Never put it in source templates, data.json, or the plain HTML build.
        key = os.environ.get("GROQ_API_KEY", "").strip()
        if not key and os.name == "nt":
            import winreg
            try:
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Environment") as registry:
                    key = str(winreg.QueryValueEx(registry, "GROQ_API_KEY")[0]).strip()
            except OSError:
                pass
        config = json.dumps({"key": key, "model": "openai/gpt-oss-120b"}).replace("<", "\\u003c")
        html = html.replace("</head>", "<script>window.TREASURY_AI=" + config + ";</script></head>", 1)
        return protect_html(html, users_path, ACCESS_DASHBOARD, ACCESS_TITLE)
    except (FileNotFoundError, ValueError) as exc:
        raise TreasuryFlowError(str(exc)) from exc


def _run_git(repo: Path, *args: str) -> str:
    result = subprocess.run(
        ["git", "-C", str(repo), *args],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
    )
    if result.returncode:
        raise RuntimeError((result.stderr or result.stdout).strip())
    return result.stdout.strip()


def publish_dashboard(index_path: Path, log: Callable[[str], None] | None = None) -> bool:
    configured = os.environ.get(PUBLISH_REPO_ENV, "").strip()
    if not configured:
        return False
    repo = Path(configured).expanduser().resolve()
    if not (repo / ".git").is_dir():
        raise TreasuryFlowError(f"پوشه انتشار GitHub معتبر نیست: {repo}")
    target = repo / "index.html"
    pending = repo / "index.html.pending"
    shutil.copy2(index_path, pending)
    os.replace(pending, target)
    _run_git(repo, "add", "-f", "--", "index.html")
    changed = subprocess.run(
        ["git", "-C", str(repo), "diff", "--cached", "--quiet", "--", "index.html"],
        timeout=30,
    ).returncode != 0
    if not changed:
        return False
    _run_git(repo, "commit", "--only", "-m", "publish protected TreasuryFlow dashboard", "--", "index.html")
    _run_git(repo, "push", "origin", "HEAD:main")
    if log:
        log("نسخه رمزگذاری‌شده داشبورد روی GitHub Pages منتشر شد.")
    return True


def process_sources(
    home: Path,
    sources: SourceSet | None = None,
    *,
    password: str = DEFAULT_PASSWORD,
    horizon_end: str = DEFAULT_HORIZON,
    annual_rate: float = 0.23,
    force: bool = False,
    log: Callable[[str], None] | None = None,
) -> ProcessResult | None:
    home = Path(home).resolve()
    home.mkdir(parents=True, exist_ok=True)
    sources = sources or find_sources(home)
    if not sources.daily:
        raise TreasuryFlowError("فایل گزارش روزانه خزانه در پوشه پیدا نشد.")

    users_path = access_file()
    if not users_path.exists():
        raise TreasuryFlowError(f"فایل کاربران پیدا نشد: {users_path}")
    signature = source_signature(sources, users_path)
    if not force and load_state(home).get("signature") == signature:
        return None

    wait_until_stable([*sources.existing(), users_path])
    report_date = detect_report_date(sources.daily, password)
    if report_date > horizon_end:
        raise TreasuryFlowError(
            f"تاریخ گزارش {report_date} بعد از پایان افق مدل ({horizon_end}) است."
        )

    work_dir = home / ".treasuryflow"
    snapshots = home / "snapshots"
    work_dir.mkdir(exist_ok=True)
    snapshots.mkdir(exist_ok=True)
    pending_data = work_dir / "data.pending.json"
    final_data = work_dir / "data.json"
    pending_data.unlink(missing_ok=True)

    env = {
        "TREASURYFLOW_SRC": str(sources.daily),
        "TREASURYFLOW_REPORT_DATE": report_date,
        "TREASURYFLOW_FACILITIES_LEDGER": str(sources.facilities or ""),
        "TREASURYFLOW_MANUAL_INPUTS": str(sources.manual or ""),
        "TREASURYFLOW_PASSWORD": password,
        "TREASURYFLOW_DATA_JSON": str(pending_data),
        "TREASURYFLOW_SNAPSHOT_DIR": str(snapshots),
        "TREASURYFLOW_HORIZON_END": horizon_end,
        "TREASURYFLOW_ANNUAL_RATE": str(annual_rate),
    }
    old_env = {key: os.environ.get(key) for key in env}
    output_buffer = io.StringIO()

    def emit(message: str) -> None:
        log_message(home, message)
        if log:
            log(message)

    emit(f"شروع پردازش گزارش {report_date}: {sources.daily.name}")
    if not sources.manual:
        emit("هشدار: فایل Manual_Inputs پیدا نشد؛ بخش‌های دستی خالی خواهند بود.")
    if not sources.facilities:
        emit("هشدار: گزارش تسهیلات پیدا نشد؛ تاریخ افتتاح چرخه‌ها تخمینی می‌شود.")

    try:
        os.environ.update(env)
        with redirect_stdout(output_buffer), redirect_stderr(output_buffer):
            runpy.run_path(
                str(resource_root() / "pipeline" / "build_forecast.py"),
                run_name="__treasuryflow_pipeline__",
            )
    finally:
        for key, old_value in old_env.items():
            if old_value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = old_value

    pipeline_log = output_buffer.getvalue().strip()
    if pipeline_log:
        for line in pipeline_log.splitlines():
            emit(line)
    if not pending_data.exists():
        raise TreasuryFlowError("پایپ‌لاین فایل داده خروجی را نساخت.")

    report_flat = report_date.replace("/", "-")
    report_path = home / f"Positron_TMS_v2_{report_flat}.html"
    plain_path = work_dir / "dashboard.pending.html"
    plain_path.unlink(missing_ok=True)
    _load_html_builder()(pending_data, plain_path)
    protected_html = _protect_html(plain_path.read_text(encoding="utf-8"), users_path)
    plain_path.unlink(missing_ok=True)
    _atomic_text(report_path, protected_html)

    index_path = home / "index.html"
    _atomic_text(index_path, protected_html)
    os.replace(pending_data, final_data)

    result = ProcessResult(
        report_date=report_date,
        report_path=report_path,
        index_path=index_path,
        data_path=final_data,
        signature=signature,
        sources=sources,
    )
    save_state(home, result)
    try:
        publish_dashboard(index_path, emit)
    except Exception as exc:
        emit(f"هشدار انتشار GitHub Pages: {exc}")
    clear_upload_request(home)
    emit(f"گزارش با موفقیت ساخته شد: {report_path.name}")
    return result


def import_files(home: Path, files: Iterable[Path], password: str = DEFAULT_PASSWORD) -> list[Path]:
    home = Path(home).resolve()
    home.mkdir(parents=True, exist_ok=True)
    imported: list[Path] = []
    for source in (Path(item).resolve() for item in files):
        if source.suffix.lower() != ".xlsx":
            raise TreasuryFlowError(f"فقط فایل Excel با پسوند xlsx پذیرفته می‌شود: {source.name}")
        kind = classify_path(source, password)
        if kind == "daily":
            report_date = detect_report_date(source, password).replace("/", "-")
            destination = home / f"Treasury Daily Report - {report_date}.xlsx"
        elif kind == "manual":
            destination = home / "Manual_Inputs.xlsx"
        else:
            destination = home / "گزارش تسهیلات.xlsx"
        if source == destination:
            imported.append(destination)
            continue
        pending = destination.with_name(destination.name + ".pending")
        shutil.copy2(source, pending)
        os.replace(pending, destination)
        imported.append(destination)
    return imported
